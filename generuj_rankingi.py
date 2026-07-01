"""
generuj_rankingi.py
───────────────────
Główny skrypt narzędzia. Czyta ustawienia z config.py, podstawia je do
zapytań w queries/*.sql, odpala na bazie i zapisuje sformatowany .xlsx
z osobnymi arkuszami.

URUCHOMIENIE:
    python generuj_rankingi.py

    # albo nadpisz wybrane parametry z linii poleceń (bez ruszania configa):
    python generuj_rankingi.py --top-n 30 --min-mecze 3 --rank-by play
    python generuj_rankingi.py --no-progres --player-group outfield

Każdy parametr z config.py ma odpowiednik w argumentach (patrz --help).

WYNIK:
    data/ranking_<liga>_<sezon>_<data>.xlsx
    Arkusze (zależnie od ustawień):
      • "Ogólny – polowi"      (jeśli MAX_PER_CLUB=None lub osobno)
      • "Max N z klubu – polowi"
      • "Progres – polowi"
      • analogiczne dla bramkarzy, jeśli PLAYER_GROUP="both"/"keeper"
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from sqlalchemy import create_engine, text

import config as cfg


# ══════════════════════════════════════════════════════════════════════════════
# ARGUMENTY (opcjonalne nadpisanie configa)
# ══════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(
        description="Generator rankingów zawodników dla WZPN.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--rank-by", choices=["league", "play"], default=cfg.RANK_BY,
                   help="Grupowanie: po kategorii (league) czy konkretnej grupie (play)")
    p.add_argument("--top-n", type=int, default=cfg.TOP_N,
                   help="Ile pozycji w rankingu")
    p.add_argument("--max-per-club", type=int, default=cfg.MAX_PER_CLUB,
                   help="Max zawodników z klubu (0 = bez limitu)")
    p.add_argument("--club-level", choices=["club", "team"], default=cfg.CLUB_LEVEL,
                   help="Limit liczony po klubie czy po drużynie")
    p.add_argument("--min-mecze", type=int, default=cfg.MIN_MECZE,
                   help="Minimalna liczba meczów do kwalifikacji")
    p.add_argument("--player-group", choices=["outfield", "keeper", "both"],
                   default=cfg.PLAYER_GROUP, help="Polowi / bramkarze / oba")
    p.add_argument("--season-id", default=cfg.SEASON_ID, help="UUID sezonu")
    p.add_argument("--no-progres", action="store_true",
                   help="Pomiń arkusz progresu")
    p.add_argument("--no-ogolny", action="store_true",
                   help="Pomiń arkusz ogólny (bez limitu klubowego)")
    return p.parse_args()


# ══════════════════════════════════════════════════════════════════════════════
# POŁĄCZENIE
# ══════════════════════════════════════════════════════════════════════════════

def make_engine():
    if not all([cfg.DB_NAME, cfg.DB_USER, cfg.DB_PASSWORD]):
        sys.exit(
            "BŁĄD: brakuje danych połączenia. Uzupełnij DB_NAME / DB_USER / "
            "DB_PASSWORD w pliku .env (patrz .env.example) lub w config.py.\n"
            "Pamiętaj: jeśli łączysz się przez SSH, najpierw zestaw tunel."
        )
    url = (f"postgresql+psycopg2://{cfg.DB_USER}:{cfg.DB_PASSWORD}"
           f"@{cfg.DB_HOST}:{cfg.DB_PORT}/{cfg.DB_NAME}")
    eng = create_engine(url, connect_args={"connect_timeout": 30})
    with eng.connect() as conn:
        conn.execute(text("SELECT 1"))
    return eng


# ══════════════════════════════════════════════════════════════════════════════
# BUDOWA ZAPYTAŃ — podstawienie placeholderów
# ══════════════════════════════════════════════════════════════════════════════

QUERIES_DIR = Path(__file__).parent / "queries"


def _league_in_clause():
    """Filtr `x.league_id IN (...)` — restrykcja po kategorii rozgrywek.
    Region NIE jest tutaj — region filtrujemy przez play_id w _play_filter_clause,
    bo ten sam league_id potrafi występować w plays różnych regionów."""
    if cfg.LEAGUE_IDS:
        return cfg.sql_in_clause(cfg.LEAGUE_IDS)
    if cfg.REGION_IDS:
        # Fallback: gdy nie podano LEAGUE_IDS, ale jest region — bierzemy
        # wszystkie league_id, które mają jakikolwiek play w tym regionie.
        # Sam region i tak zostanie wyegzekwowany w _play_filter_clause.
        return ("SELECT DISTINCT league_id FROM plays "
                f"WHERE region_id IN ({cfg.sql_in_clause(cfg.REGION_IDS)})")
    return cfg.sql_in_clause([])


def _play_filter_clause():
    """Dodatkowe filtry: jawne PLAY_IDS i/lub region przez plays.region_id.
    KLUCZOWE: region filtrujemy po play_id (NIE league_id), bo ten sam
    league_id bywa współdzielony między regionami."""
    conditions = []
    if cfg.PLAY_IDS:
        conditions.append(f"x.play_id IN ({cfg.sql_in_clause(cfg.PLAY_IDS)})")
    if cfg.REGION_IDS:
        conditions.append(
            "x.play_id IN ("
            f"SELECT _id FROM plays WHERE region_id IN ({cfg.sql_in_clause(cfg.REGION_IDS)})"
            ")"
        )
    if not conditions:
        return ""
    return "AND " + " AND ".join(conditions)


def _common_placeholders(args, keeper_value, max_per_club_effective):
    """Placeholdery wspólne dla ranking.sql i progres.sql."""
    by_play = args.rank_by == "play"

    if by_play:
        join_plays      = "JOIN plays d ON s.play_id = d._id"
        play_name_sel   = "d.name"
        rank_group_col  = "d.name"
        rank_group_fin  = "play_name"
    else:
        join_plays      = ""
        play_name_sel   = "'Ranking Ogólny'"
        rank_group_col  = "c.name"
        rank_group_fin  = "league_name"

    club_partition = "cl._id" if args.club_level == "club" else "b._id"

    return {
        "season_id":            args.season_id,
        "league_in":            _league_in_clause(),
        "play_filter":          _play_filter_clause(),
        "keeper_value":         "true" if keeper_value else "false",
        "min_mecze":            args.min_mecze,
        "join_plays":           join_plays,
        "play_name_select":     play_name_sel,
        "rank_group_col":       rank_group_col,
        "rank_group_col_final": rank_group_fin,
        "club_partition":       club_partition,
        "max_per_club":         max_per_club_effective,
        "top_n":                args.top_n,
    }


def load_query(filename, placeholders):
    """Wczytuje plik .sql i podstawia placeholdery {nazwa}.

    Uwaga: używamy ręcznego podstawiania zamiast str.format(), bo SQL/komentarze
    mogą zawierać przypadkowe nawiasy klamrowe (np. '{...}' w komentarzu albo
    operatory JSON), które wywaliłyby .format() błędem IndexError/KeyError.
    Podstawiamy tylko znane, jawnie zadeklarowane placeholdery.
    """
    sql = (QUERIES_DIR / filename).read_text(encoding="utf-8")
    for key, value in placeholders.items():
        sql = sql.replace("{" + key + "}", str(value))
    # kontrola: czy zostały niepodstawione, znane-wyglądające placeholdery?
    import re
    leftover = re.findall(r"\{(season_id|league_in|rank_group_col|"
                          r"rank_group_col_final|join_plays|keeper_value|"
                          r"min_mecze|club_partition|max_per_club|top_n|"
                          r"play_name_select|play_filter|a_start|a_end|"
                          r"b_start|b_end|progres_min)\}", sql)
    if leftover:
        raise ValueError(
            f"Niepodstawione placeholdery w {filename}: {sorted(set(leftover))}. "
            f"Sprawdź, czy wszystkie są przekazane w _common_placeholders()."
        )
    return sql


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — formatowanie w stylu build_reports.py
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT_HEADER = Font(name=cfg.XLSX_FONT_NAME, size=10, bold=True, color=cfg.COLOR_HEADER_TXT)
FONT_BODY   = Font(name=cfg.XLSX_FONT_NAME, size=10)
FONT_BOLD   = Font(name=cfg.XLSX_FONT_NAME, size=10, bold=True)
FILL_HEADER = PatternFill("solid", start_color=cfg.COLOR_HEADER_BG)
FILL_TOP3   = PatternFill("solid", start_color=cfg.COLOR_TOP3_BG)
ALIGN_C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Mapa: nazwa kolumny z SQL → (nagłówek w Excelu, szerokość, wyrównanie)
COLMAP = {
    "rank_position": ("#",            5,  ALIGN_C),
    "firstname":     ("Imię",         14, ALIGN_L),
    "lastname":      ("Nazwisko",     18, ALIGN_L),
    "yob":           ("Rocznik",      8,  ALIGN_C),
    "club_name":     ("Klub",         26, ALIGN_L),
    "team_name":     ("Drużyna",      26, ALIGN_L),
    "league_name":   ("Kategoria",    20, ALIGN_L),
    "play_name":     ("Liga (grupa)", 24, ALIGN_L),
    "avg_score":     ("Śr. ocena",    10, ALIGN_C),
    "mecze_count":   ("Mecze",        7,  ALIGN_C),
    "avg_a":         ("Śr. okres A",  11, ALIGN_C),
    "avg_b":         ("Śr. okres B",  11, ALIGN_C),
    "progres":       ("Progres",      10, ALIGN_C),
    "mecze_a":       ("Mecze A",      8,  ALIGN_C),
    "mecze_b":       ("Mecze B",      8,  ALIGN_C),
    "player_id":     ("player_id",    16, ALIGN_L),
}


def write_sheet(wb, title, df, first_sheet=False):
    """Zapisuje DataFrame jako sformatowany arkusz. Resetuje rank_position
    osobno per grupa (kategoria/play), żeby numeracja '#' była czytelna."""
    if first_sheet:
        ws = wb.active
        ws.title = title[:31]   # Excel limit 31 znaków
    else:
        ws = wb.create_sheet(title[:31])

    if df.empty:
        ws.cell(row=1, column=1, value="Brak danych dla tych ustawień.").font = FONT_BODY
        return

    cols = [c for c in df.columns if c in COLMAP]
    headers = [COLMAP[c][0] for c in cols]

    # Nagłówek
    ws.append(headers)
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C

    # Wiersze
    for _, row in df.iterrows():
        excel_row = [row[c] for c in cols]
        ws.append(excel_row)
        r = ws.max_row
        # podświetl TOP 3 każdej grupy
        is_top3 = int(row.get("rank_position", 99)) <= 3
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=j)
            cell.font = FONT_BOLD if (c == "lastname" and is_top3) else FONT_BODY
            cell.alignment = COLMAP[c][2]
            if is_top3:
                cell.fill = FILL_TOP3

    # Szerokości
    for j, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = COLMAP[c][1]

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# GŁÓWNA LOGIKA
# ══════════════════════════════════════════════════════════════════════════════

def fetch(engine, sql):
    with engine.connect() as conn:
        return pd.read_sql(text(sql), conn)


def _group_col(df):
    """Wybiera kolumnę grupującą: play_name jeśli rankujemy po play
    (czyli wartości są realnymi nazwami grup, nie 'Ranking Ogólny'),
    w przeciwnym razie league_name."""
    if "play_name" in df.columns and df["play_name"].iloc[0] != "Ranking Ogólny":
        return "play_name"
    return "league_name"


def renumber_per_group(df, sort_col="avg_score"):
    """Po pobraniu danych numeruje pozycje od 1 w obrębie każdej grupy
    (kategorii albo play). sort_col = kolumna, po której malejąco sortujemy
    (avg_score dla zwykłego rankingu, progres dla progresu)."""
    if df.empty:
        return df
    if sort_col not in df.columns:
        # awaryjnie: pierwsza dostępna kolumna liczbowa wyniku
        sort_col = "progres" if "progres" in df.columns else "avg_score"
    gcol = _group_col(df)
    df = df.sort_values([gcol, sort_col], ascending=[True, False]).copy()
    df["rank_position"] = df.groupby(gcol).cumcount() + 1
    return df


def build_for_group(engine, args, keeper_value, label, wb, state):
    """Generuje arkusze (ogólny / max-klub / progres) dla jednej grupy
    (polowi albo bramkarze)."""
    # --- limit klubowy ---
    # 999999 = praktycznie bez limitu
    max_club = 999999 if (args.max_per_club in (None, 0)) else args.max_per_club

    # 1) Ranking OGÓLNY (bez limitu klubowego) — jeśli chcemy
    if not args.no_ogolny:
        ph = _common_placeholders(args, keeper_value, 999999)
        sql = load_query("ranking.sql", ph)
        df = renumber_per_group(fetch(engine, sql))
        write_sheet(wb, f"Ogólny – {label}", df, first_sheet=state["first"])
        state["first"] = False
        print(f"  [{label}] Ogólny: {len(df)} wierszy")

    # 2) Ranking MAX N z klubu — jeśli ustawiono limit
    if args.max_per_club not in (None, 0):
        ph = _common_placeholders(args, keeper_value, max_club)
        sql = load_query("ranking.sql", ph)
        df = renumber_per_group(fetch(engine, sql))
        write_sheet(wb, f"Max {args.max_per_club} z klubu – {label}", df,
                    first_sheet=state["first"])
        state["first"] = False
        print(f"  [{label}] Max {args.max_per_club}/klub: {len(df)} wierszy")

    # 3) Progres
    if cfg.GENERATE_PROGRES and not args.no_progres:
        ph = _common_placeholders(args, keeper_value, max_club)
        ph.update({
            "a_start": cfg.PROGRES_OKRES_A[0], "a_end": cfg.PROGRES_OKRES_A[1],
            "b_start": cfg.PROGRES_OKRES_B[0], "b_end": cfg.PROGRES_OKRES_B[1],
            "progres_min": cfg.PROGRES_MIN_MECZE,
        })
        sql = load_query("progres.sql", ph)
        df = renumber_per_group(fetch(engine, sql), sort_col="progres")
        write_sheet(wb, f"Progres – {label}", df, first_sheet=state["first"])
        state["first"] = False
        print(f"  [{label}] Progres: {len(df)} wierszy")


def main():
    args = parse_args()
    print("Łączenie z bazą...")
    engine = make_engine()
    print("OK\n")

    wb = Workbook()
    state = {"first": True}

    groups = []
    if args.player_group in ("outfield", "both"):
        groups.append((False, "polowi"))
    if args.player_group in ("keeper", "both"):
        groups.append((True, "bramkarze"))

    print(f"Generuję rankingi (rank_by={args.rank_by}, top_n={args.top_n}, "
          f"min_mecze={args.min_mecze}, max/klub={args.max_per_club})...")
    for keeper_value, label in groups:
        build_for_group(engine, args, keeper_value, label, wb, state)

    # Nazwa pliku
    Path(cfg.OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = Path(cfg.OUTPUT_DIR) / f"ranking_{args.rank_by}_{stamp}.xlsx"
    wb.save(out)
    print(f"\nZapisano: {out}")


if __name__ == "__main__":
    main()
